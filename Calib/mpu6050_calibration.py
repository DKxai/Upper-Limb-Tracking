#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MPU6050 calibration toolkit (numpy-only).

Hai quy trình hiệu chuẩn, theo 2 bài báo:

  GYRO  — Wang et al., "In-field gyroscope autocalibration with iterative
          attitude estimation", Mechatronics 102 (2024) 103232.
          Mô hình:  g = K ⊙ (m + b)   (K = diag scale, b = bias)
          Tham chiếu: xoay tay MỖI trục đúng 360°. Thuật toán lặp ước lượng
          thái độ (Rodrigues) sao cho góc tích phân = 360°.

  ACCEL — Hassan et al., "A Field Calibration Method for Low-Cost MEMS
          Accelerometer Based on the Generalized Nonlinear Least Square Method".
          Mô hình:  a = S (ã − B)   (S đối xứng 3x3, B bias 3-vector)
          Tham chiếu: ||a|| = g ở mọi tư thế tĩnh. Giải bằng GNLS/LM.

Phụ thuộc: numpy, pandas, openpyxl   (KHÔNG cần scipy)

Cách dùng:
  python mpu6050_calibration.py gyro-template  [gyro_template.xlsx]
  python mpu6050_calibration.py accel-template [accel_template.xlsx]
  python mpu6050_calibration.py gyro  <data.xlsx> [report.xlsx]
  python mpu6050_calibration.py accel <data.xlsx> [report.xlsx]
  python mpu6050_calibration.py validate-gyro  <calib.xlsx> <verify.xlsx> [out.xlsx]
  python mpu6050_calibration.py validate-accel <calib.xlsx> <verify.xlsx> [out.xlsx]
  python mpu6050_calibration.py node <N> <gyro.xlsx> <accel.xlsx> [out_dir]  # calib 1 trong 6 node
  python mpu6050_calibration.py build-js [dir]  # gộp node*_calib.json -> khối CALIBRATION dán vào JS
  python mpu6050_calibration.py demo            # sinh dữ liệu mẫu + chạy cả 2, in báo cáo
"""

import sys
import os
import glob
import json
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill

# MPU6050: hệ số đổi raw-count -> đơn vị vật lý (đặt theo FS_SEL/AFS_SEL của bạn)
GYRO_LSB_PER_DPS = 131.0      # ±250 °/s
ACCEL_LSB_PER_G = 16384.0     # ±2 g
LOCAL_G = 9.80665             # m/s^2

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True)


# ════════════════════════════════════════════════════════════════════════
#   THUẬT TOÁN DÙNG CHUNG (numpy)
# ════════════════════════════════════════════════════════════════════════
def _rodrigues(u, theta):
    """Ma trận quay quanh trục đơn vị u một góc theta (rad) — Eq.(4) bài gyro."""
    ux = np.array([[0, -u[2], u[1]], [u[2], 0, -u[0]], [-u[1], u[0], 0]])
    return (np.cos(theta) * np.eye(3) + np.sin(theta) * ux
            + (1 - np.cos(theta)) * np.outer(u, u))


def _newton(F, x0, iters=60, tol=1e-12):
    """Newton đa biến với Jacobian sai-phân + damping (giải F(x)=0)."""
    x = np.asarray(x0, float).copy()
    for _ in range(iters):
        f = F(x)
        if np.linalg.norm(f) < tol:
            break
        J = np.zeros((len(f), len(x)))
        for j in range(len(x)):
            h = 1e-6 * max(1.0, abs(x[j]))
            dx = np.zeros_like(x); dx[j] = h
            J[:, j] = (F(x + dx) - f) / h
        try:
            step = np.linalg.solve(J, -f)
        except np.linalg.LinAlgError:
            step = np.linalg.lstsq(J, -f, rcond=None)[0]
        a = 1.0
        for _ in range(25):
            if np.linalg.norm(F(x + a * step)) < np.linalg.norm(f):
                break
            a *= 0.5
        x = x + a * step
    return x


# ════════════════════════════════════════════════════════════════════════
#   GYRO — Paper 1 (Wang et al. 2024)
# ════════════════════════════════════════════════════════════════════════
def calibrate_gyro(M_dps, stage, fs):
    """
    M_dps : (N,3) gyro thô (°/s)
    stage : (N,) 0=đứng yên, 1=xoay X, 2=xoay Y, 3=xoay Z (đều 360°)
    fs    : tần số lấy mẫu (Hz)
    Trả về: K(3), bias(3, °/s)  với mô hình  g = K ⊙ (m + bias)
    """
    M_dps = np.asarray(M_dps, float)
    stage = np.asarray(stage, int)
    dt = 1.0 / fs
    theta_total = 360.0  # độ

    # Eq.(2): bias = -trung bình giai đoạn đứng yên (để g = K(m+b) = 0 khi đứng yên)
    m0 = M_dps[stage == 0]
    if len(m0) < 5:
        raise ValueError("Thiếu dữ liệu giai đoạn đứng yên (stage=0).")
    bias = -m0.mean(axis=0)

    rot_stages = [s for s in (1, 2, 3) if np.any(stage == s)]
    if not rot_stages:
        raise ValueError("Không thấy giai đoạn xoay (stage=1/2/3).")

    def build_A_T(K):
        """Với K hiện tại: tích phân thái độ, dựng A_s (Eq.8-10) và thời lượng T_s."""
        A_list, T_list = [], []
        for s in rot_stages:
            idx = np.where(stage == s)[0]
            Mhat = M_dps[idx] + bias            # m̂ đã bù bias (°/s)
            T = len(idx) * dt
            C = np.eye(3)
            A = np.zeros((3, 3))
            for t in range(len(idx)):
                w_rad = np.deg2rad(K * Mhat[t])  # g = K·m̂ -> rad/s
                ang = np.linalg.norm(w_rad) * dt  # θ_t (rad)
                A += C @ np.diag(Mhat[t])         # C̃ = C·diag(m̂);  Ḡ = (1/N)ΣC̃·K
                if ang > 1e-12:
                    u = w_rad / np.linalg.norm(w_rad)
                    C = C @ _rodrigues(u, ang)
            A_list.append(A / len(idx))
            T_list.append(T)
        return A_list, T_list

    K = np.ones(3)
    for _ in range(60):
        A_list, T_list = build_A_T(K)

        def Fsys(k):
            # Eq.(11): ||A_s k||^2 = (θ_total/T_s)^2  cho từng giai đoạn xoay
            return np.array([(A_list[i] @ k) @ (A_list[i] @ k)
                             - (theta_total / T_list[i]) ** 2
                             for i in range(len(rot_stages))])

        if len(rot_stages) == 3:
            Knew = _newton(Fsys, K)
        else:
            # < 3 trục: chỉ tinh chỉnh theo trục trội mỗi giai đoạn (ít gặp)
            Knew = K.copy()
            for i, s in enumerate(rot_stages):
                ax = np.argmax(np.abs(M_dps[stage == s]).mean(axis=0))
                target = theta_total / T_list[i]
                cur = np.linalg.norm(A_list[i] @ Knew)
                if cur > 1e-9:
                    Knew[ax] *= target / cur
        if np.linalg.norm(Knew - K) <= 1e-5:
            K = Knew
            break
        K = Knew
    return K, bias


def _gyro_total_angle(M_dps, bias, K, fs):
    """
    Góc QUAY THỰC tích lũy (độ) theo thời gian = ‖∫ C·ω dt‖ trong khung ban đầu.
    Đây ĐÚNG là đại lượng calib tối ưu (‖Ḡ‖·T → 360°). KHÁC với "độ dài đường"
    cumsum‖ω‖dt (bản cũ) vốn bị thổi phồng bởi rung tay / nhiễu / thời gian giữ yên
    → khiến chart nhìn như calib "không đổi". Khớp đúng tích phân trong build_A_T.
    """
    dt = 1.0 / fs
    g = K * (M_dps + bias)                  # °/s sau khi áp K,bias (raw nếu K=1,bias=0)
    w = np.deg2rad(g)
    wn = np.linalg.norm(w, axis=1)
    C = np.eye(3)
    A = np.zeros(3)                          # vector quay tích lũy trong khung ban đầu (độ)
    out = np.empty(len(M_dps))
    for t in range(len(M_dps)):
        A = A + C @ g[t] * dt               # cộng ω (khung ban đầu) TRƯỚC khi xoay C
        out[t] = np.linalg.norm(A)
        if wn[t] * dt > 1e-12:
            C = C @ _rodrigues(w[t] / wn[t], wn[t] * dt)
    return out


# ════════════════════════════════════════════════════════════════════════
#   ACCEL — Paper 2 (Hassan et al., GNLS)
# ════════════════════════════════════════════════════════════════════════
def _S_from_p(p):
    return np.array([[p[0], p[3], p[4]],
                     [p[3], p[1], p[5]],
                     [p[4], p[5], p[2]]])


def calibrate_accel(A_g, g_ref=1.0):
    """
    A_g  : (N,3) accel thô (đơn vị g), mỗi hàng = 1 tư thế tĩnh.
    g_ref: độ lớn trọng trường mục tiêu (1.0 nếu dữ liệu theo g).
    Trả về: S(3x3 đối xứng), B(3) — mô hình a = S(ã − B).  Giải bằng LM (họ GNLS).
    """
    A_g = np.asarray(A_g, float)

    def resid(p):
        S, b = _S_from_p(p), p[6:9]
        ac = (A_g - b) @ S.T            # mỗi hàng: S(ã−B)
        return np.sum(ac * ac, axis=1) - g_ref ** 2   # ||a||^2 − g^2

    p = np.array([1, 1, 1, 0, 0, 0, 0, 0, 0], float)
    lam = 1e-3
    prev = np.sum(resid(p) ** 2)
    for _ in range(300):
        r = resid(p)
        # Jacobian sai-phân
        J = np.zeros((len(r), 9))
        for j in range(9):
            h = 1e-7 * max(1.0, abs(p[j]))
            dp = np.zeros(9); dp[j] = h
            J[:, j] = (resid(p + dp) - r) / h
        H = J.T @ J
        grad = J.T @ r
        try:
            step = np.linalg.solve(H + lam * np.diag(np.diag(H) + 1e-12), -grad)
        except np.linalg.LinAlgError:
            step = np.linalg.lstsq(H + lam * np.eye(9), -grad, rcond=None)[0]
        pn = p + step
        cost = np.sum(resid(pn) ** 2)
        if cost < prev:                 # nhận bước -> giảm damping (Levenberg-Marquardt)
            p, prev, lam = pn, cost, lam * 0.7
            if np.linalg.norm(step) < 1e-12:
                break
        else:
            lam *= 2.5
            if lam > 1e12:
                break
    return _S_from_p(p), p[6:9]


def _accel_norm(A_g, S, B):
    """||a|| (g) sau khi a = S(ã−B)."""
    ac = (A_g - B) @ S.T
    return np.linalg.norm(ac, axis=1)


# ════════════════════════════════════════════════════════════════════════
#   EXCEL — TEMPLATE
# ════════════════════════════════════════════════════════════════════════
def _style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def make_gyro_template(path, demo=False, fs=100):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    cols = ["time_s", "stage", "gx_dps", "gy_dps", "gz_dps"]
    ws.append(cols)
    _style_header(ws, 1, len(cols))

    if demo:
        df = _synth_gyro(fs)
        for _, r in df.iterrows():
            ws.append([round(r.time_s, 3), int(r.stage),
                       round(r.gx_dps, 4), round(r.gy_dps, 4), round(r.gz_dps, 4)])

    ins = wb.create_sheet("HuongDan")
    text = [
        ["HIỆU CHUẨN GYRO — quy trình xoay 360° (Wang et al. 2024)"],
        [""],
        [f"• Tần số lấy mẫu fs = {fs} Hz (đổi GYRO_LSB_PER_DPS trong script nếu khác FS_SEL)."],
        ["• gx/gy/gz: vận tốc góc THÔ, đơn vị °/s = raw_count / 131 (FS=±250°/s)."],
        ["• Cột 'stage' đánh dấu giai đoạn:"],
        ["     0 = ĐỨNG YÊN tuyệt đối (3-5 giây) — để ước lượng bias."],
        ["     1 = xoay tay MỘT vòng 360° quanh trục X rồi về vị trí cũ."],
        ["     2 = xoay 360° quanh trục Y."],
        ["     3 = xoay 360° quanh trục Z."],
        ["• Giữa các lần xoay nên dừng vài giây; xoay ĐỀU TAY, một chiều."],
        ["• Mỗi giai đoạn chỉ cần ~3-5s. Không cần thiết bị chuẩn nào."],
        [""],
        ["Sau khi điền xong, chạy:  python mpu6050_calibration.py gyro <file.xlsx>"],
    ]
    for row in text:
        ins.append(row)
    ins["A1"].font = Font(bold=True, size=13)

    wb.save(path)
    return path


def make_accel_template(path, demo=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    cols = ["position", "ax_g", "ay_g", "az_g"]
    ws.append(cols)
    _style_header(ws, 1, len(cols))

    if demo:
        df = _synth_accel()
        for i, r in df.iterrows():
            ws.append([i + 1, round(r.ax_g, 5), round(r.ay_g, 5), round(r.az_g, 5)])

    ins = wb.create_sheet("HuongDan")
    text = [
        ["HIỆU CHUẨN ACCEL — GNLS / ||a||=g (Hassan et al.)"],
        [""],
        ["• ax/ay/az: gia tốc TĨNH, đơn vị g = raw_count / 16384 (AFS=±2g)."],
        ["• Mỗi HÀNG = 1 tư thế tĩnh. Giữ cảm biến ĐỨNG YÊN, ghi vài giây rồi"],
        ["  lấy TRUNG BÌNH cho ra 1 hàng (giảm nhiễu)."],
        ["• Thu ≥ 12 (lý tưởng ~30) tư thế phủ đều mặt cầu:"],
        ["  6 mặt (±X, ±Y, ±Z hướng lên) + các nghiêng ~45° quanh X và Y."],
        ["• Chỉ trọng lực được cảm nhận (không rung, không di chuyển)."],
        [""],
        ["Sau khi điền xong, chạy:  python mpu6050_calibration.py accel <file.xlsx>"],
    ]
    for row in text:
        ins.append(row)
    ins["A1"].font = Font(bold=True, size=13)

    wb.save(path)
    return path


# ════════════════════════════════════════════════════════════════════════
#   EXCEL — REPORT (có chart so sánh trước/sau)
# ════════════════════════════════════════════════════════════════════════
def _write_table(ws, start_row, headers, rows):
    ws.append([]) if False else None
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=1 + j, value=h)
    _style_header(ws, start_row, len(headers))
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            ws.cell(row=start_row + 1 + i, column=1 + j, value=v)
    return start_row + 1 + len(rows)


def report_gyro(M_dps, stage, fs, K, bias, out_path):
    wb = Workbook()
    summ = wb.active
    summ.title = "TongHop"

    summ["A1"] = "KẾT QUẢ HIỆU CHUẨN GYRO (mô hình g = K ⊙ (m + bias))"
    summ["A1"].font = Font(bold=True, size=13)
    nxt = _write_table(summ, 3, ["Trục", "Scale K", "Bias (°/s)"],
                       [["X", round(K[0], 6), round(bias[0], 6)],
                        ["Y", round(K[1], 6), round(bias[1], 6)],
                        ["Z", round(K[2], 6), round(bias[2], 6)]])

    # So sánh góc tích phân mỗi trục: thô vs đã hiệu chuẩn vs 360°
    axis_names = {1: "X", 2: "Y", 3: "Z"}
    final_rows = []
    for s in (1, 2, 3):
        if not np.any(stage == s):
            continue
        idx = np.where(stage == s)[0]
        Ms = M_dps[idx]
        raw = _gyro_total_angle(Ms, np.zeros(3), np.ones(3), fs)
        cal = _gyro_total_angle(Ms, bias, K, fs)
        t = np.arange(len(idx)) / fs

        sh = wb.create_sheet(f"Xoay_{axis_names[s]}")
        sh.append(["time_s", "Goc_THO_deg", "Goc_HIEUCHUAN_deg", "Chuan_360"])
        _style_header(sh, 1, 4)
        for i in range(len(idx)):
            sh.append([round(t[i], 3), round(raw[i], 3), round(cal[i], 3), 360])
        ch = LineChart()
        ch.title = f"Góc tích phân khi xoay quanh {axis_names[s]} — Thô vs Hiệu chuẩn vs 360°"
        ch.x_axis.title = "Thời gian (s)"
        ch.y_axis.title = "Góc tích lũy (°)"
        ch.height, ch.width = 9, 18
        data = Reference(sh, min_col=2, max_col=4, min_row=1, max_row=len(idx) + 1)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(Reference(sh, min_col=1, min_row=2, max_row=len(idx) + 1))
        sh.add_chart(ch, "F2")

        final_rows.append([axis_names[s], round(raw[-1], 2), round(cal[-1], 2),
                           round(abs(raw[-1] - 360), 2), round(abs(cal[-1] - 360), 2)])

    nxt = _write_table(summ, nxt + 2,
                       ["Trục", "Góc cuối THÔ (°)", "Góc cuối HC (°)",
                        "|Sai số| THÔ (°)", "|Sai số| HC (°)"], final_rows)

    # Bar chart sai số trước/sau
    bsh = wb.create_sheet("SaiSo")
    bsh.append(["Trục", "Sai so THO (deg)", "Sai so HIEUCHUAN (deg)"])
    _style_header(bsh, 1, 3)
    for r in final_rows:
        bsh.append([r[0], r[3], r[4]])
    bc = BarChart()
    bc.title = "Sai số góc so với 360° — Trước vs Sau hiệu chuẩn"
    bc.y_axis.title = "|Góc − 360| (°)"
    bc.height, bc.width = 9, 14
    bc.add_data(Reference(bsh, min_col=2, max_col=3, min_row=1, max_row=len(final_rows) + 1),
                titles_from_data=True)
    bc.set_categories(Reference(bsh, min_col=1, min_row=2, max_row=len(final_rows) + 1))
    bsh.add_chart(bc, "E2")

    # Snippet dán vào CalibrationData.js (dự án dùng g = K*(g_raw + bias))
    snip = wb.create_sheet("ChepVaoCode")
    snip["A1"] = "Dán vào js/utils/CalibrationData.js -> gyro (K, bias cùng quy ước g=K*(g_raw+bias)):"
    snip["A1"].font = Font(bold=True)
    snip["A3"] = (f"gyro: {{ K: [{K[0]:.8f}, {K[1]:.8f}, {K[2]:.8f}], "
                  f"bias: [{bias[0]:.8f}, {bias[1]:.8f}, {bias[2]:.8f}] }}")
    snip["A3"].font = Font(name="Consolas")

    wb.save(out_path)
    return out_path


def report_accel(A_g, S, B, out_path, g_ref=1.0):
    wb = Workbook()
    summ = wb.active
    summ.title = "TongHop"
    summ["A1"] = "KẾT QUẢ HIỆU CHUẨN ACCEL (mô hình a = S (ã − B))"
    summ["A1"].font = Font(bold=True, size=13)

    nxt = _write_table(summ, 3, ["S (ma trận đối xứng)", "col X", "col Y", "col Z"],
                       [["row X", round(S[0, 0], 6), round(S[0, 1], 6), round(S[0, 2], 6)],
                        ["row Y", round(S[1, 0], 6), round(S[1, 1], 6), round(S[1, 2], 6)],
                        ["row Z", round(S[2, 0], 6), round(S[2, 1], 6), round(S[2, 2], 6)]])
    nxt = _write_table(summ, nxt + 1, ["Bias B", "X", "Y", "Z"],
                       [["g", round(B[0], 6), round(B[1], 6), round(B[2], 6)],
                        ["m/s²", round(B[0] * LOCAL_G, 6), round(B[1] * LOCAL_G, 6),
                         round(B[2] * LOCAL_G, 6)]])

    norm_raw = np.linalg.norm(A_g, axis=1)
    norm_cal = _accel_norm(A_g, S, B)
    err_raw = norm_raw - g_ref
    err_cal = norm_cal - g_ref
    rms = lambda x: float(np.sqrt(np.mean(x ** 2)))
    nxt = _write_table(summ, nxt + 1,
                       ["Chỉ tiêu (||a|| so với g)", "TRƯỚC", "SAU"],
                       [["RMS sai số (g)", round(rms(err_raw), 5), round(rms(err_cal), 5)],
                        ["Max |sai số| (g)", round(float(np.max(np.abs(err_raw))), 5),
                         round(float(np.max(np.abs(err_cal))), 5)],
                        ["Std ||a|| (g)", round(float(np.std(norm_raw)), 5),
                         round(float(np.std(norm_cal)), 5)]])

    # Chart: ||a|| theo từng tư thế — thô vs hiệu chuẩn vs chuẩn g
    cmp = wb.create_sheet("DoiChieu")
    cmp.append(["position", "norm_THO_g", "norm_HIEUCHUAN_g", "chuan_g"])
    _style_header(cmp, 1, 4)
    n = len(A_g)
    for i in range(n):
        cmp.append([i + 1, round(float(norm_raw[i]), 5),
                    round(float(norm_cal[i]), 5), g_ref])
    ch = LineChart()
    ch.title = "Độ lớn ||a|| mỗi tư thế — Thô vs Hiệu chuẩn vs chuẩn g (lý tưởng = g)"
    ch.x_axis.title = "Tư thế #"
    ch.y_axis.title = "||a|| (g)"
    ch.height, ch.width = 9, 18
    ch.add_data(Reference(cmp, min_col=2, max_col=4, min_row=1, max_row=n + 1),
                titles_from_data=True)
    ch.set_categories(Reference(cmp, min_col=1, min_row=2, max_row=n + 1))
    cmp.add_chart(ch, "F2")

    snip = wb.create_sheet("ChepVaoCode")
    snip["A1"] = "Dán vào js/utils/CalibrationData.js -> accel (M = S, bias đơn vị m/s²):"
    snip["A1"].font = Font(bold=True)
    M = S
    snip["A3"] = ("accel: { M: [["
                  f"{M[0,0]:.6f}, {M[0,1]:.6f}, {M[0,2]:.6f}], ["
                  f"{M[1,0]:.6f}, {M[1,1]:.6f}, {M[1,2]:.6f}], ["
                  f"{M[2,0]:.6f}, {M[2,1]:.6f}, {M[2,2]:.6f}]], "
                  f"bias: [{B[0]*LOCAL_G:.6f}, {B[1]*LOCAL_G:.6f}, {B[2]*LOCAL_G:.6f}] }}")
    snip["A3"].font = Font(name="Consolas")

    wb.save(out_path)
    return out_path


# ════════════════════════════════════════════════════════════════════════
#   DỮ LIỆU MÔ PHỎNG (cho demo + điền template mẫu)
# ════════════════════════════════════════════════════════════════════════
def _synth_gyro(fs=100, seed=0):
    rng = np.random.default_rng(seed)
    K_true = np.array([1.05, 0.97, 1.02])
    b_true = np.array([0.8, -0.6, 1.1])          # °/s (mô hình g=K(m+b) -> bias = b_true)
    dt = 1.0 / fs
    rows = []
    t = 0.0

    def push(g_vec, st):
        nonlocal t
        # m = g/K_true − b_true  + nhiễu  (đảo mô hình để calib hồi phục đúng K,b)
        m = g_vec / K_true - b_true + rng.normal(0, 0.15, 3)
        rows.append([t, st, m[0], m[1], m[2]])
        t += dt

    for _ in range(int(3 * fs)):                 # S0: đứng yên 3s
        push(np.zeros(3), 0)
    T = 4.0
    for ax, st in [(0, 1), (1, 2), (2, 3)]:      # xoay 360° quanh X,Y,Z
        n = int(T * fs)
        wmax = 360.0 * np.pi / (2 * T)           # ∫ wmax·sin(πt/T) dt = 360°
        for i in range(n):
            g = np.zeros(3)
            g[ax] = wmax * np.sin(np.pi * i / n)
            push(g, st)
        for _ in range(int(1.5 * fs)):           # nghỉ giữa các lần
            push(np.zeros(3), 0)
    return pd.DataFrame(rows, columns=["time_s", "stage", "gx_dps", "gy_dps", "gz_dps"])


def _synth_accel(seed=1):
    rng = np.random.default_rng(seed)
    S_true = np.array([[1.012, 0.010, -0.008],
                       [0.010, 0.991, 0.006],
                       [-0.008, 0.006, 1.005]])
    B_true = np.array([0.020, -0.030, 0.015])    # g
    Sinv = np.linalg.inv(S_true)
    rows = []
    # 6 mặt + nghiêng ~45° quanh X và Y => ~30 tư thế
    base = []
    for s in (+1, -1):
        base += [[s, 0, 0], [0, s, 0], [0, 0, s]]
    tilts = []
    for v in base:
        v = np.array(v, float)
        for ang in (np.deg2rad(45), np.deg2rad(-45)):
            R = _rodrigues(np.array([1.0, 0, 0]), ang)
            tilts.append(R @ v)
            R = _rodrigues(np.array([0, 1.0, 0]), ang)
            tilts.append(R @ v)
    poses = base + tilts
    for a in poses:
        a = np.array(a, float)
        a = a / np.linalg.norm(a)                # |a_true| = 1 g
        raw = Sinv @ a + B_true + rng.normal(0, 0.004, 3)   # ã = S^{-1}a + B + nhiễu
        rows.append([raw[0], raw[1], raw[2]])
    return pd.DataFrame(rows, columns=["ax_g", "ay_g", "az_g"])


# ════════════════════════════════════════════════════════════════════════
#   CLI
# ════════════════════════════════════════════════════════════════════════
def _read(path, sheet="Data"):
    return pd.read_excel(path, sheet_name=sheet)


def run_gyro(data_path, out_path=None, fs=100):
    df = _read(data_path)
    M = df[["gx_dps", "gy_dps", "gz_dps"]].values
    stage = df["stage"].values
    K, bias = calibrate_gyro(M, stage, fs)
    out_path = out_path or os.path.splitext(data_path)[0] + "_report.xlsx"
    report_gyro(M, np.asarray(stage, int), fs, K, bias, out_path)
    print(f"[GYRO] K = {np.round(K,5)}  bias(°/s) = {np.round(bias,5)}")
    print(f"[GYRO] Báo cáo: {out_path}")
    return K, bias


def run_accel(data_path, out_path=None):
    df = _read(data_path)
    A = df[["ax_g", "ay_g", "az_g"]].values
    S, B = calibrate_accel(A, g_ref=1.0)
    out_path = out_path or os.path.splitext(data_path)[0] + "_report.xlsx"
    report_accel(A, S, B, out_path, g_ref=1.0)
    print(f"[ACCEL] S =\n{np.round(S,5)}\n[ACCEL] B(g) = {np.round(B,5)}")
    print(f"[ACCEL] Báo cáo: {out_path}")
    return S, B


def validate_gyro(calib_path, verify_path, out_path=None, fs=100):
    """Fit K,bias trên tập CALIB → đo Trước/Sau trên tập VERIFY (held-out, xoay 360° mới)."""
    dfc = _read(calib_path)
    K, bias = calibrate_gyro(dfc[["gx_dps", "gy_dps", "gz_dps"]].values,
                             dfc["stage"].values, fs)
    dfv = _read(verify_path)
    Mv = dfv[["gx_dps", "gy_dps", "gz_dps"]].values
    stagev = np.asarray(dfv["stage"].values, int)
    out_path = out_path or os.path.splitext(verify_path)[0] + "_truocsau.xlsx"
    report_gyro(Mv, stagev, fs, K, bias, out_path)   # chart/bảng Thô vs Hiệu chuẩn vs 360°
    print(f"[VALIDATE GYRO] Tham số fit trên {os.path.basename(calib_path)}: "
          f"K={np.round(K,5)} bias={np.round(bias,5)}")
    print(f"[VALIDATE GYRO] Báo cáo Trước/Sau (trên {os.path.basename(verify_path)}): {out_path}")


def validate_accel(calib_path, verify_path, out_path=None):
    """Fit S,B trên tập CALIB → đo Trước/Sau ‖a‖ trên tập VERIFY (held-out, tư thế mới)."""
    dfc = _read(calib_path)
    S, B = calibrate_accel(dfc[["ax_g", "ay_g", "az_g"]].values, g_ref=1.0)
    dfv = _read(verify_path)
    Av = dfv[["ax_g", "ay_g", "az_g"]].values
    out_path = out_path or os.path.splitext(verify_path)[0] + "_truocsau.xlsx"
    report_accel(Av, S, B, out_path, g_ref=1.0)      # chart ‖a‖ + bảng RMS/Max Trước–Sau
    print(f"[VALIDATE ACCEL] Tham số fit trên {os.path.basename(calib_path)}")
    print(f"[VALIDATE ACCEL] Báo cáo Trước/Sau (trên {os.path.basename(verify_path)}): {out_path}")


# ════════════════════════════════════════════════════════════════════════
#   NHIỀU NODE — calib 6 con MPU6050 rồi ghép thành CALIBRATION cho web app
# ════════════════════════════════════════════════════════════════════════
def _entry_js(node, e, indent="  "):
    """Một entry CALIBRATION[node] đúng định dạng CalibrationData.js."""
    M, ab = e["accel"]["M"], e["accel"]["bias_ms2"]
    K, gb = e["gyro"]["K"], e["gyro"]["bias"]
    return (
        f"{indent}{node}: {{\n"
        f"{indent}  accel: {{\n"
        f"{indent}    M: [\n"
        f"{indent}      [{M[0][0]:.6f}, {M[0][1]:.6f}, {M[0][2]:.6f}],\n"
        f"{indent}      [{M[1][0]:.6f}, {M[1][1]:.6f}, {M[1][2]:.6f}],\n"
        f"{indent}      [{M[2][0]:.6f}, {M[2][1]:.6f}, {M[2][2]:.6f}]\n"
        f"{indent}    ],\n"
        f"{indent}    bias: [{ab[0]:.6f}, {ab[1]:.6f}, {ab[2]:.6f}]  // m/s²\n"
        f"{indent}  }},\n"
        f"{indent}  gyro: {{\n"
        f"{indent}    K: [{K[0]:.8f}, {K[1]:.8f}, {K[2]:.8f}],\n"
        f"{indent}    bias: [{gb[0]:.8f}, {gb[1]:.8f}, {gb[2]:.8f}]  // °/s\n"
        f"{indent}  }}\n"
        f"{indent}}}"
    )


def run_node(node_id, gyro_path, accel_path, out_dir=None, fs=100):
    """Calib 1 node: gyro + accel → 2 báo cáo + nodeN_calib.json (để build-js gộp lại)."""
    out_dir = out_dir or os.path.dirname(os.path.abspath(gyro_path)) or "."
    dfg = _read(gyro_path)
    K, gbias = calibrate_gyro(dfg[["gx_dps", "gy_dps", "gz_dps"]].values,
                              np.asarray(dfg["stage"].values, int), fs)
    report_gyro(dfg[["gx_dps", "gy_dps", "gz_dps"]].values,
                np.asarray(dfg["stage"].values, int), fs, K, gbias,
                os.path.join(out_dir, f"node{node_id}_gyro_report.xlsx"))

    dfa = _read(accel_path)
    S, B = calibrate_accel(dfa[["ax_g", "ay_g", "az_g"]].values, g_ref=1.0)
    report_accel(dfa[["ax_g", "ay_g", "az_g"]].values, S, B,
                 os.path.join(out_dir, f"node{node_id}_accel_report.xlsx"), g_ref=1.0)

    entry = {
        "node": int(node_id),
        "gyro": {"K": K.tolist(), "bias": gbias.tolist()},
        "accel": {"M": S.tolist(), "bias_ms2": (B * LOCAL_G).tolist()},
    }
    jpath = os.path.join(out_dir, f"node{node_id}_calib.json")
    with open(jpath, "w", encoding="utf-8") as f:
        json.dump(entry, f, ensure_ascii=False, indent=2)
    print(f"[NODE {node_id}] gyro K={np.round(K,4)} bias={np.round(gbias,4)}")
    print(f"[NODE {node_id}] accel S diag={np.round(np.diag(S),4)} bias(m/s²)={np.round(B*LOCAL_G,4)}")
    print(f"[NODE {node_id}] -> {jpath}")
    return entry


def build_js(src, out_path=None):
    """Gộp các node*_calib.json (thư mục hoặc danh sách file) → khối CALIBRATION dán vào JS."""
    if isinstance(src, str) and os.path.isdir(src):
        files = sorted(glob.glob(os.path.join(src, "node*_calib.json")))
        base_dir = src
    else:
        files = list(src) if not isinstance(src, str) else [src]
        base_dir = os.path.dirname(os.path.abspath(files[0])) if files else "."
    entries = {}
    for fp in files:
        with open(fp, encoding="utf-8") as f:
            e = json.load(f)
        entries[int(e["node"])] = e
    if not entries:
        print("Không thấy file node*_calib.json nào.")
        return

    lines = ["// === Dán đè khối CALIBRATION trong js/utils/CalibrationData.js ===",
             "export const CALIBRATION = {"]
    for node in range(1, 7):
        if node in entries:
            lines.append(_entry_js(node, entries[node]) + ",")
        else:
            lines.append(f"  {node}: DEFAULT_CALIB,  // CHƯA calib — đang dùng mặc định")
    lines.append("};")
    block = "\n".join(lines)

    out_path = out_path or os.path.join(base_dir, "CALIBRATION_nodes.js")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(block + "\n")
    done = sorted(entries.keys())
    missing = [n for n in range(1, 7) if n not in entries]
    print(f"[BUILD-JS] Đã calib node: {done}" + (f" | CÒN THIẾU: {missing}" if missing else " | đủ 6/6"))
    print(f"[BUILD-JS] Khối JS: {out_path}\n")
    print(block)
    return out_path


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # console Windows in được tiếng Việt
    except Exception:
        pass
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        return
    cmd = args[0]
    here = os.path.dirname(os.path.abspath(__file__))

    if cmd == "gyro-template":
        out = args[1] if len(args) > 1 else os.path.join(here, "gyro_template.xlsx")
        print("Đã tạo template gyro (kèm dữ liệu mẫu):", make_gyro_template(out, demo=True))

    elif cmd == "accel-template":
        out = args[1] if len(args) > 1 else os.path.join(here, "accel_template.xlsx")
        print("Đã tạo template accel (kèm dữ liệu mẫu):", make_accel_template(out, demo=True))

    elif cmd == "gyro":
        if len(args) < 2:
            print("Thiếu file dữ liệu. Vd: python mpu6050_calibration.py gyro gyro_template.xlsx")
            return
        run_gyro(args[1], args[2] if len(args) > 2 else None)

    elif cmd == "accel":
        if len(args) < 2:
            print("Thiếu file dữ liệu. Vd: python mpu6050_calibration.py accel accel_template.xlsx")
            return
        run_accel(args[1], args[2] if len(args) > 2 else None)

    elif cmd == "validate-gyro":
        if len(args) < 3:
            print("Cách dùng: python mpu6050_calibration.py validate-gyro <calib.xlsx> <verify.xlsx> [out.xlsx]")
            return
        validate_gyro(args[1], args[2], args[3] if len(args) > 3 else None)

    elif cmd == "validate-accel":
        if len(args) < 3:
            print("Cách dùng: python mpu6050_calibration.py validate-accel <calib.xlsx> <verify.xlsx> [out.xlsx]")
            return
        validate_accel(args[1], args[2], args[3] if len(args) > 3 else None)

    elif cmd == "node":
        if len(args) < 4:
            print("Cách dùng: python mpu6050_calibration.py node <N> <gyro.xlsx> <accel.xlsx> [out_dir]")
            return
        run_node(int(args[1]), args[2], args[3], args[4] if len(args) > 4 else None)

    elif cmd == "build-js":
        src = args[1] if len(args) > 1 else here
        build_js(src, args[2] if len(args) > 2 else None)

    elif cmd == "demo":
        gt = make_gyro_template(os.path.join(here, "gyro_template.xlsx"), demo=True)
        at = make_accel_template(os.path.join(here, "accel_template.xlsx"), demo=True)
        print("Template:", gt, at)
        K, b = run_gyro(gt, os.path.join(here, "gyro_report.xlsx"))
        S, B = run_accel(at, os.path.join(here, "accel_report.xlsx"))
        print("\n--- Kiểm chứng demo (hồi phục tham số mô phỏng) ---")
        print("Gyro K_true=[1.05,0.97,1.02] b_true=[0.8,-0.6,1.1]")
        print("Accel S_true diag≈[1.012,0.991,1.005] B_true=[0.02,-0.03,0.015]")
    else:
        print(__doc__)


if __name__ == "__main__":
    main()
